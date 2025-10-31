"""
Servicio de exportación a Excel
backend/app/services/export_service.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List
from datetime import datetime
from sqlalchemy.orm import Session

from app.models.miembro import Miembro
from app.models.pago import Pago
from app.models.categoria import Categoria


class ExportService:
    """Servicio para exportar datos a Excel"""
    
    @staticmethod
    def _apply_header_style(worksheet, row_num: int = 1):
        """Aplicar estilo a los headers de la tabla"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border
    
    @staticmethod
    def _auto_adjust_columns(worksheet):
        """Ajustar automáticamente el ancho de las columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _add_metadata(worksheet, title: str):
        """Agregar metadata al inicio del documento"""
        # Título
        worksheet.insert_rows(1, 2)
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14, color="366092")
        
        # Fecha de generación
        worksheet['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        worksheet['A2'].font = Font(italic=True, size=10, color="666666")
        
        # Ajustar fila de headers
        worksheet.row_dimensions[3].height = 25
    
    @staticmethod
    def generar_excel_socios(socios: List[Miembro]) -> bytes:
        """
        Generar archivo Excel con listado de socios
        
        Args:
            socios: Lista de objetos Miembro
        
        Returns:
            bytes: Archivo Excel en formato bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Socios"
        
        # Headers
        headers = [
            "N° Socio",
            "DNI",
            "Nombre Completo",
            "Email",
            "Teléfono",
            "Categoría",
            "Cuota Mensual",
            "Estado",
            "Fecha Alta",
            "Deuda Actual",
            "Días Mora"
        ]
        ws.append(headers)
        
        # Datos
        for socio in socios:
            ws.append([
                socio.numero_miembro,
                socio.dni,
                socio.nombre_completo,
                socio.email or "Sin email",
                socio.telefono or "Sin teléfono",
                socio.categoria.nombre if socio.categoria else "Sin categoría",
                f"${socio.categoria.cuota_mensual:,.2f}" if socio.categoria else "$0.00",
                socio.estado,
                socio.fecha_alta.strftime('%d/%m/%Y') if socio.fecha_alta else "",
                f"${socio.deuda_actual:,.2f}" if hasattr(socio, 'deuda_actual') else "$0.00",
                socio.dias_mora if hasattr(socio, 'dias_mora') else 0
            ])
        
        # Aplicar estilos
        ExportService._apply_header_style(ws)
        ExportService._auto_adjust_columns(ws)
        ExportService._add_metadata(ws, "Listado de Socios")
        
        # Congelar primera fila (headers)
        ws.freeze_panes = "A4"
        
        # Guardar en BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def generar_excel_pagos(pagos: List[Pago], db: Session) -> bytes:
        """
        Generar archivo Excel con listado de pagos
        
        Args:
            pagos: Lista de objetos Pago
            db: Sesión de base de datos
        
        Returns:
            bytes: Archivo Excel en formato bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        
        # Headers
        headers = [
            "ID Pago",
            "Fecha",
            "N° Socio",
            "Nombre Socio",
            "Concepto",
            "Monto",
            "Método Pago",
            "Estado",
            "Mes/Año Pago",
            "Registrado Por",
            "Observaciones"
        ]
        ws.append(headers)
        
        # Datos
        total_monto = 0
        for pago in pagos:
            # Cargar relaciones si es necesario
            if hasattr(pago, 'miembro') and pago.miembro:
                nombre_socio = pago.miembro.nombre_completo
                numero_socio = pago.miembro.numero_miembro
            else:
                nombre_socio = "N/A"
                numero_socio = "N/A"
            
            registrado_por = "Sistema"
            if hasattr(pago, 'usuario') and pago.usuario:
                registrado_por = pago.usuario.username
            
            mes_anio = ""
            if pago.mes and pago.anio:
                mes_anio = f"{pago.mes:02d}/{pago.anio}"
            
            ws.append([
                pago.id,
                pago.fecha_pago.strftime('%d/%m/%Y %H:%M') if pago.fecha_pago else "",
                numero_socio,
                nombre_socio,
                pago.concepto or "Cuota mensual",
                f"${pago.monto:,.2f}",
                pago.metodo_pago,
                pago.estado,
                mes_anio,
                registrado_por,
                pago.observaciones or ""
            ])
            
            if pago.estado != "ANULADO":
                total_monto += pago.monto
        
        # Fila de total
        ws.append([])
        total_row = ws.max_row
        ws[f'E{total_row}'] = "TOTAL:"
        ws[f'E{total_row}'].font = Font(bold=True)
        ws[f'F{total_row}'] = f"${total_monto:,.2f}"
        ws[f'F{total_row}'].font = Font(bold=True, color="006600")
        
        # Aplicar estilos
        ExportService._apply_header_style(ws)
        ExportService._auto_adjust_columns(ws)
        ExportService._add_metadata(ws, "Listado de Pagos")
        
        # Congelar primera fila
        ws.freeze_panes = "A4"
        
        # Guardar en BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def generar_excel_morosidad(morosos: List[dict]) -> bytes:
        """
        Generar archivo Excel con reporte de morosidad
        
        Args:
            morosos: Lista de diccionarios con datos de morosos
        
        Returns:
            bytes: Archivo Excel en formato bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Morosidad"
        
        # Headers
        headers = [
            "N° Socio",
            "DNI",
            "Nombre Completo",
            "Email",
            "Teléfono",
            "Categoría",
            "Cuota Mensual",
            "Deuda Total",
            "Días de Mora",
            "Último Pago",
            "Estado"
        ]
        ws.append(headers)
        
        # Datos
        total_deuda = 0
        for moroso in morosos:
            ws.append([
                moroso.get("numero_miembro", ""),
                moroso.get("dni", ""),
                moroso.get("nombre_completo", ""),
                moroso.get("email", "Sin email"),
                moroso.get("telefono", "Sin teléfono"),
                moroso.get("categoria", "Sin categoría"),
                f"${moroso.get('cuota_mensual', 0):,.2f}",
                f"${moroso.get('deuda', 0):,.2f}",
                moroso.get("dias_mora", 0),
                moroso.get("ultimo_pago", "Nunca"),
                moroso.get("estado", "MOROSO")
            ])
            total_deuda += moroso.get("deuda", 0)
        
        # Fila de totales
        ws.append([])
        summary_row = ws.max_row
        ws[f'A{summary_row}'] = "RESUMEN:"
        ws[f'A{summary_row}'].font = Font(bold=True, size=12)
        
        ws.append([])
        ws.append([
            "",
            "",
            "",
            "",
            "",
            "Total Socios Morosos:",
            len(morosos),
            "Deuda Total:",
            f"${total_deuda:,.2f}",
            "",
            ""
        ])
        
        summary_data_row = ws.max_row
        ws[f'F{summary_data_row}'].font = Font(bold=True)
        ws[f'H{summary_data_row}'].font = Font(bold=True)
        ws[f'I{summary_data_row}'].font = Font(bold=True, color="CC0000")
        
        # Aplicar estilos
        ExportService._apply_header_style(ws)
        ExportService._auto_adjust_columns(ws)
        ExportService._add_metadata(ws, "Reporte de Morosidad")
        
        # Congelar primera fila
        ws.freeze_panes = "A4"
        
        # Resaltar filas según días de mora
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        orange_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        for row in range(4, ws.max_row - 2):  # Excluir headers y resumen
            dias_mora_cell = ws[f'I{row}']
            try:
                dias_mora = int(dias_mora_cell.value) if dias_mora_cell.value else 0
                if dias_mora > 30:
                    for col in range(1, 12):
                        ws.cell(row=row, column=col).fill = red_fill
                elif dias_mora > 15:
                    for col in range(1, 12):
                        ws.cell(row=row, column=col).fill = orange_fill
            except:
                pass
        
        # Guardar en BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def generar_excel_accesos(accesos: List[dict]) -> bytes:
        """
        Generar archivo Excel con historial de accesos
        
        Args:
            accesos: Lista de diccionarios con datos de accesos
        
        Returns:
            bytes: Archivo Excel en formato bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Accesos"
        
        # Headers
        headers = [
            "ID",
            "Fecha/Hora",
            "N° Socio",
            "Nombre Socio",
            "Tipo Acceso",
            "Resultado",
            "Ubicación",
            "Dispositivo",
            "Observaciones"
        ]
        ws.append(headers)
        
        # Contadores
        total_exitosos = 0
        total_denegados = 0
        
        # Datos
        for acceso in accesos:
            resultado = acceso.get("permitido", False)
            if resultado:
                total_exitosos += 1
            else:
                total_denegados += 1
            
            ws.append([
                acceso.get("id", ""),
                acceso.get("fecha_hora", ""),
                acceso.get("numero_miembro", ""),
                acceso.get("nombre_socio", ""),
                acceso.get("tipo_acceso", "QR"),
                "EXITOSO" if resultado else "DENEGADO",
                acceso.get("ubicacion", "Principal"),
                acceso.get("dispositivo_id", "N/A"),
                acceso.get("motivo_denegacion", "") if not resultado else ""
            ])
        
        # Resumen
        ws.append([])
        summary_row = ws.max_row
        ws[f'A{summary_row}'] = "ESTADÍSTICAS:"
        ws[f'A{summary_row}'].font = Font(bold=True, size=12)
        
        ws.append([
            "",
            "Total Accesos:",
            len(accesos),
            "",
            "Exitosos:",
            total_exitosos,
            "",
            "Denegados:",
            total_denegados
        ])
        
        stats_row = ws.max_row
        ws[f'B{stats_row}'].font = Font(bold=True)
        ws[f'E{stats_row}'].font = Font(bold=True)
        ws[f'F{stats_row}'].font = Font(color="006600")
        ws[f'H{stats_row}'].font = Font(bold=True)
        ws[f'I{stats_row}'].font = Font(color="CC0000")
        
        # Aplicar estilos
        ExportService._apply_header_style(ws)
        ExportService._auto_adjust_columns(ws)
        ExportService._add_metadata(ws, "Historial de Accesos")
        
        # Congelar primera fila
        ws.freeze_panes = "A4"
        
        # Colorear según resultado
        green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        red_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        
        for row in range(4, ws.max_row - 1):
            resultado_cell = ws[f'F{row}']
            if resultado_cell.value == "EXITOSO":
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = green_fill
            elif resultado_cell.value == "DENEGADO":
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = red_fill
        
        # Guardar en BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()